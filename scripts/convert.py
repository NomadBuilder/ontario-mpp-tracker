#!/usr/bin/env python3
"""Convert the MPP Excel spreadsheet to JSON for the web showcase."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Current Ontario MPPs & How They Vote..xlsx"
OUTPUT = ROOT / "data" / "mpps.json"
PHOTOS = ROOT / "data" / "photos.json"
EXPENSES = ROOT / "data" / "expenses.json"
SUNSHINE = ROOT / "data" / "sunshine.json"
OLA_BILL_BASE = "https://www.ola.org/en/legislative-business/bills/parliament-44/session-1"

# Featured bills shown in filters / campaign presets — overridden by Display Settings
# rows like "Bill 5" | Yes/No when that tab is present.
DEFAULT_FEATURED = {"Bill 5", "Bill 17", "Bill 24", "Bill 60", "Bill 68", "Bill 97", "Bill 110"}

# Site display toggles — overridden by a "Display Settings" sheet tab when present.
# Yes/No (or true/false/1/0/show/hide) in column B. Unknown fields are ignored.
DEFAULT_DISPLAY = {
    "salary": False,
    "benefits": False,
    "votingAlignment": False,
    "expenses": True,
}

DISPLAY_ALIASES = {
    "salary": "salary",
    "benefits": "benefits",
    "benefit": "benefits",
    "votingalignment": "votingAlignment",
    "voting alignment": "votingAlignment",
    "alignment": "votingAlignment",
    "party alignment": "votingAlignment",
    "expenses": "expenses",
    "expense": "expenses",
    "expense disclosure": "expenses",
    "expensedisclosure": "expenses",
}


def sort_bill_ids(bills) -> list:
    def key(x: str):
        m = re.match(r"Bill\s+(\d+)$", x, re.I)
        return (0, int(m.group(1))) if m else (1, x.lower())

    return sorted(bills, key=key)


def parse_bill_field(field_raw: str) -> str | None:
    """Map sheet Field cells like 'Bill 5' / 'bill5' → 'Bill 5'."""
    if not field_raw:
        return None
    m = re.match(r"^bill\s*(\d+)\s*$", field_raw.strip(), re.I)
    if m:
        return f"Bill {m.group(1)}"
    return None


def get_xlsx_path() -> Path:
    import argparse
    parser = argparse.ArgumentParser(description="Convert MPP spreadsheet to JSON")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Path to Excel file (default: project root spreadsheet)",
    )
    args, _ = parser.parse_known_args()
    return args.xlsx if args.xlsx else XLSX


def slug_from_name(name: str) -> str:
    name = re.sub(r"^(Hon\.|Dr\.)\s*", "", name or "")
    slug = name.lower().strip()
    slug = re.sub(r"[^a-z0-9\s-]", "", slug)
    return re.sub(r"\s+", "-", slug)


def simplify_bill_name(name: str) -> str:
    if name and name.startswith("Bill "):
        parts = name.split(" ")
        return f"{parts[0]} {parts[1]}"
    return name


def normalize_bill_url(raw, bill_header: str) -> str | None:
    if not raw:
        return None
    if isinstance(raw, str) and raw.startswith("http"):
        return raw.split("?")[0]
    # Bill 110 row has title text instead of URL
    m = re.search(r"Bill\s+(\d+)", bill_header or "")
    if m:
        return f"{OLA_BILL_BASE}/bill-{m.group(1)}"
    return None


def vote_display(vote) -> dict:
    if vote == "Aye":
        return {"label": "Yes", "yes": True}
    if vote == "Nay":
        return {"label": "No", "yes": False}
    if vote == "No Show":
        return {"label": "No Show", "yes": None}
    return {"label": "N/A", "yes": None}


def load_photos() -> tuple[dict, dict]:
    if not PHOTOS.exists():
        return {}, {}
    data = json.loads(PHOTOS.read_text())
    return data.get("bySlug", {}), data.get("byName", {})


def load_expenses() -> tuple[dict, dict, str | None]:
    if not EXPENSES.exists():
        return {}, {}, None
    data = json.loads(EXPENSES.read_text())
    return data.get("bySlug", {}), data.get("byName", {}), data.get("fetchedAt")


def load_sunshine() -> tuple[dict, dict, dict | None]:
    if not SUNSHINE.exists():
        return {}, {}, None
    data = json.loads(SUNSHINE.read_text())
    meta = {
        "fetchedAt": data.get("fetchedAt"),
        "year": data.get("year"),
        "priorYear": data.get("priorYear"),
        "source": data.get("source"),
        "matched": data.get("matched"),
        "unmatched": data.get("unmatched") or [],
        "discrepancyCount": len(data.get("discrepancies") or []),
    }
    return data.get("bySlug", {}), data.get("byName", {}), meta


def lookup_sunshine(name: str, by_slug: dict, by_name: dict) -> dict | None:
    """Attach official PSSD salary row when available."""
    if not by_slug:
        return None

    slug = slug_from_name(name)
    info = by_slug.get(slug)

    if not info:
        key = re.sub(r"^(Hon\.|Dr\.)\s*", "", name or "").lower().strip()
        mapped = by_name.get(key) or by_name.get(re.sub(r"\s+", " ", key))
        if mapped:
            info = by_slug.get(mapped)

    if not info:
        parts = re.sub(r"^(Hon\.|Dr\.)\s*", "", name or "").lower().strip().split()
        if len(parts) >= 2:
            short_slug = f"{parts[0]}-{parts[-1]}"
            info = by_slug.get(short_slug)
            if not info:
                mapped = by_name.get(f"{parts[0]} {parts[-1]}")
                if mapped:
                    info = by_slug.get(mapped)

    return info


def apply_sunshine_pay(sheet_salary, sheet_benefits, sheet_as_of, sheet_raise, sunshine: dict | None):
    """Prefer official PSSD figures when matched; keep sheet as fallback.

    Returns (salary, benefits, asOf, raisePct, sunshinePayload, notes).
    """
    if not sunshine:
        return sheet_salary, sheet_benefits, sheet_as_of, sheet_raise, None, []

    notes = []
    salary = sunshine.get("salary")
    benefits = sunshine.get("benefits")
    as_of = sunshine.get("year")
    raise_pct = sunshine.get("raisePct")

    if sheet_salary is not None and salary is not None and abs(float(sheet_salary) - float(salary)) > 1:
        notes.append(f"salary sheet={sheet_salary} pssd={salary}")
    if sheet_benefits is not None and benefits is not None and abs(float(sheet_benefits) - float(benefits)) > 1:
        notes.append(f"benefits sheet={sheet_benefits} pssd={benefits}")
    if sheet_salary is None and salary is not None:
        notes.append(f"filled salary from pssd={salary}")
    if (sheet_benefits is None or float(sheet_benefits or 0) == 0) and benefits and float(benefits) > 0:
        if sheet_benefits is None:
            notes.append(f"filled benefits from pssd={benefits}")
        elif abs(float(sheet_benefits) - float(benefits)) > 1:
            pass  # already noted as mismatch above

    payload = {
        "salary": salary,
        "benefits": benefits,
        "jobTitle": sunshine.get("jobTitle"),
        "year": sunshine.get("year"),
        "priorSalary": sunshine.get("priorSalary"),
        "priorYear": sunshine.get("priorYear"),
        "raisePct": raise_pct,
        "listedAs": f"{sunshine.get('firstName')} {sunshine.get('lastName')}".strip(),
        "source": sunshine.get("source"),
    }
    # Prefer PSSD (authoritative public disclosure); fall back to sheet if somehow empty
    return (
        salary if salary is not None else sheet_salary,
        benefits if benefits is not None else sheet_benefits,
        as_of if as_of is not None else sheet_as_of,
        raise_pct if raise_pct is not None else sheet_raise,
        payload,
        notes,
    )


def lookup_photo(name: str, by_slug: dict, by_name: dict) -> tuple[str | None, str | None]:
    slug = slug_from_name(name)
    if slug in by_slug and by_slug[slug].get("photo"):
        info = by_slug[slug]
        return info["photo"], info.get("profileUrl")

    key = re.sub(r"^(Hon\.|Dr\.)\s*", "", name).lower().strip()
    if key in by_name:
        return by_name[key], f"https://www.ola.org/en/members/all/{slug}"

    parts = key.split()
    if len(parts) > 2:
        short_slug = f"{parts[0]}-{parts[-1]}"
        if short_slug in by_slug and by_slug[short_slug].get("photo"):
            info = by_slug[short_slug]
            return info["photo"], info.get("profileUrl")

    return None, f"https://www.ola.org/en/members/all/{slug}"


def lookup_expenses(name: str, by_slug: dict, by_name: dict, fetched_at: str | None) -> dict | None:
    """Attach OLA expense disclosure summary when available."""
    if not by_slug:
        return None

    slug = slug_from_name(name)
    info = by_slug.get(slug)

    if not info:
        key = re.sub(r"^(Hon\.|Dr\.)\s*", "", name or "").lower().strip()
        mapped = by_name.get(key)
        if mapped:
            info = by_slug.get(mapped)

    if not info:
        parts = re.sub(r"^(Hon\.|Dr\.)\s*", "", name or "").lower().strip().split()
        if len(parts) >= 2:
            short_slug = f"{parts[0]}-{parts[-1]}"
            info = by_slug.get(short_slug)
            if not info:
                mapped = by_name.get(f"{parts[0]} {parts[-1]}")
                if mapped:
                    info = by_slug.get(mapped)

    if not info:
        return None

    return {
        "total": info.get("total", 0),
        "travel": info.get("travel", 0),
        "accommodation": info.get("accommodation", 0),
        "meals": info.get("meals", 0),
        "hospitality": info.get("hospitality", 0),
        "claimCount": info.get("claimCount", 0),
        "periodCount": info.get("periodCount", 0),
        "sourceUrl": info.get("sourceUrl"),
        "asOf": fetched_at,
    }


def parse_show_value(raw) -> bool | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    s = str(raw).strip().lower()
    if s in {"yes", "y", "true", "1", "show", "on"}:
        return True
    if s in {"no", "n", "false", "0", "hide", "off"}:
        return False
    return None


def find_display_settings_sheet(wb) -> str | None:
    """Match common Display Settings tab names (Site Display, OAC · Display Settings, …)."""
    exact = {"display settings", "display", "site settings", "site display"}
    for name in wb.sheetnames:
        n = name.strip().lower()
        if n in exact:
            return name
        # Title-ish tabs: "OAC · Display Settings", "Display settings (public)"
        if "display" in n and ("setting" in n or n.startswith("site ")):
            return name
    return None


def load_display_settings(wb) -> tuple[dict, set]:
    """Read optional 'Display Settings' tab: Field | Show (Yes/No).

    Field rows:
      - salary / benefits / votingAlignment / expenses — card/table visibility
      - Bill N — which bills appear in featured filters / campaign chips

    Bill rows overlay defaults: No removes a default bill, Yes adds any bill
    that exists in the vote data. If no Bill rows are present, defaults apply.
    """
    display = dict(DEFAULT_DISPLAY)
    featured = set(DEFAULT_FEATURED)
    sheet_name = find_display_settings_sheet(wb)
    if not sheet_name:
        print("No Display Settings tab — using defaults:", display, "bills:", sort_bill_ids(featured))
        return display, featured

    ws = wb[sheet_name]
    seen_bill_row = False
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if not row or not row[0]:
            continue
        field_raw = str(row[0]).strip()
        key_norm = re.sub(r"\s+", " ", field_raw).lower()
        # Skip title / blurb / header rows
        if key_norm in {"field", "setting", "name", "column"}:
            continue
        if "controls what the public" in key_norm or key_norm.startswith("oac"):
            continue
        if len(field_raw) > 80:
            continue

        show = parse_show_value(row[1] if len(row) > 1 else None)
        if show is None:
            continue

        bill = parse_bill_field(field_raw)
        if bill:
            seen_bill_row = True
            if show:
                featured.add(bill)
            else:
                featured.discard(bill)
            continue

        key = DISPLAY_ALIASES.get(key_norm) or DISPLAY_ALIASES.get(key_norm.replace(" ", ""))
        if key:
            display[key] = show

    if not seen_bill_row:
        featured = set(DEFAULT_FEATURED)

    print(f"Display Settings ({sheet_name}): {display}")
    print(f"Featured bills: {sort_bill_ids(featured)}")
    return display, featured


def find_header_row(ws) -> int:
    """Locate the row that starts with 'MPP Name' (handles inserted date rows)."""
    for r in range(1, 20):
        val = ws.cell(r, 1).value
        if val and "MPP" in str(val) and "Name" in str(val):
            return r
    raise ValueError("Could not find header row in 'How They Voted' sheet")


def find_bill_start_col(ws, header_row: int) -> int:
    """First column after Party that looks like a bill/vote header."""
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if not h:
            continue
        s = str(h).strip()
        if s.startswith("Bill ") or "Surveillance" in s or "Pricing" in s:
            return c
    return 11


def main() -> None:
    xlsx_path = get_xlsx_path()
    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    by_slug, by_name = load_photos()
    exp_by_slug, exp_by_name, exp_fetched = load_expenses()
    sun_by_slug, sun_by_name, sun_meta = load_sunshine()
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    display, featured = load_display_settings(wb)

    # Optional contact/roles sheet
    mpp_info = {}
    if "Current Ontarios MPPs" in wb.sheetnames:
        contact = wb["Current Ontarios MPPs"]
        headers = {
            (contact.cell(1, c).value or "").strip().lower(): c
            for c in range(1, contact.max_column + 1)
            if contact.cell(1, c).value
        }
        name_c = headers.get("mpp name", 1)
        roles_c = headers.get("roles")
        party_c = headers.get("party")
        riding_c = headers.get("riding")
        email_c = headers.get("email")
        phone_c = headers.get("constituency office number")
        score_c = headers.get("oac score")

        for row in contact.iter_rows(min_row=2, values_only=False):
            name_cell = row[name_c - 1].value
            if not name_cell:
                continue
            name = str(name_cell).strip()

            def cell(col):
                return row[col - 1].value if col else None

            mpp_info[name.lower()] = {
                "roles": (cell(roles_c) or "") if roles_c else "",
                "party": (cell(party_c) or "") if party_c else "",
                "riding": (cell(riding_c) or "") if riding_c else "",
                "email": (cell(email_c) or "") if email_c else "",
                "phone": (cell(phone_c) or "") if phone_c else "",
                "oacScore": cell(score_c) if score_c is not None and cell(score_c) is not None else 0,
            }

    ws = wb["How They Voted"]
    header_row = find_header_row(ws)
    bill_start = find_bill_start_col(ws, header_row)
    print(f"Header row: {header_row}, bill columns start at: {bill_start}")

    bill_headers = []
    bill_urls: dict[str, str | None] = {}
    for i in range(bill_start, ws.max_column + 1):
        h = ws.cell(header_row, i).value
        if not h or not str(h).strip():
            continue
        h = str(h).strip()
        # Skip non-bill helper columns
        if h.lower() in {"party", "url", "third reading vote", "first reading"}:
            continue
        bill_headers.append((i, h))
        simple = simplify_bill_name(h)
        raw_url = ws.cell(1, i).value
        url = normalize_bill_url(raw_url, h)
        bill_urls[simple] = url
        bill_urls[h] = url

    print(f"Bills found ({len(bill_headers)}): {[h for _, h in bill_headers]}")

    data_start = header_row + 1

    party_votes: dict = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row[4]:
            continue
        party = row[9]
        if party not in party_votes:
            party_votes[party] = {h: {"Aye": 0, "Nay": 0} for _, h in bill_headers}
        for col_idx, bill_name in bill_headers:
            vote = row[col_idx - 1]
            if vote in ("Aye", "Nay"):
                party_votes[party][bill_name][vote] += 1

    party_majority = {}
    for party, bills in party_votes.items():
        party_majority[party] = {}
        for bill, counts in bills.items():
            party_majority[party][bill] = "Aye" if counts["Aye"] >= counts["Nay"] else "Nay"

    def find_extra(name: str, email: str) -> dict | None:
        for key, info in mpp_info.items():
            if key in name.lower() or name.lower() in key:
                return info
        for info in mpp_info.values():
            if info.get("email") and email and info["email"].lower() == email.lower():
                return info
        return None

    mpps = []
    skipped = 0
    sunshine_notes: list[str] = []
    with_sunshine = 0
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row[4]:
            skipped += 1
            continue

        name = str(row[4]).strip()
        # Skip accidental header/date leftovers
        if name.lower() in {"name", "mpp name"} or not row[1]:
            skipped += 1
            continue

        party = row[9] or ""
        email = row[3] or ""
        extra = find_extra(name, email)
        photo, profile_url = lookup_photo(name, by_slug, by_name)
        expenses = lookup_expenses(name, exp_by_slug, exp_by_name, exp_fetched)
        sunshine_row = lookup_sunshine(name, sun_by_slug, sun_by_name)
        salary, benefits, as_of, raise_pct, sunshine_payload, notes = apply_sunshine_pay(
            row[5], row[6], row[7], row[8], sunshine_row
        )
        if sunshine_payload:
            with_sunshine += 1
        for n in notes:
            sunshine_notes.append(f"{name}: {n}")

        votes = []
        aligned = total = 0
        for col_idx, bill_name in bill_headers:
            vote = row[col_idx - 1]
            simple = simplify_bill_name(bill_name)
            if vote in ("Aye", "Nay"):
                total += 1
                majority = party_majority.get(party, {}).get(bill_name)
                if majority and vote == majority:
                    aligned += 1
            vd = vote_display(vote)
            votes.append({
                "bill": simple,
                "billFull": bill_name.strip(),
                "url": bill_urls.get(simple) or bill_urls.get(bill_name.strip()),
                "vote": vote,
                "display": vd["label"],
                "yes": vd["yes"],
            })

        mpps.append({
            "name": name,
            "lastName": row[1] or "",
            "firstName": row[2] or "",
            "email": email,
            "party": party,
            "salary": salary,
            "benefits": benefits,
            "asOf": as_of,
            "raisePct": raise_pct,
            "riding": extra["riding"] if extra else "",
            "roles": extra["roles"] if extra else "",
            "phone": (str(extra["phone"]).strip() if extra and extra["phone"] else ""),
            "oacScore": extra["oacScore"] if extra else 0,
            "photo": photo,
            "profileUrl": profile_url,
            "expenses": expenses,
            "sunshine": sunshine_payload,
            "votingAlignment": round(aligned / total * 100) if total else None,
            "votes": votes,
        })

    mpps.sort(key=lambda m: str(m["lastName"]).lower())

    bills_meta = []
    seen = set()
    for _, h in bill_headers:
        simple = simplify_bill_name(h)
        if simple in seen:
            continue
        seen.add(simple)
        bills_meta.append({
            "id": simple,
            "label": h.strip(),
            "url": bill_urls.get(simple),
            "featured": simple in featured,
        })

    # Only advertise featured bills that actually appear in the vote columns
    available = {b["id"] for b in bills_meta}
    featured_list = sort_bill_ids(b for b in featured if b in available)

    OUTPUT.parent.mkdir(exist_ok=True)
    payload = {
        "display": display,
        "featuredBills": featured_list,
        "bills": bills_meta,
        "mpps": mpps,
    }
    if sun_meta:
        payload["sunshine"] = {
            "fetchedAt": sun_meta.get("fetchedAt"),
            "year": sun_meta.get("year"),
            "priorYear": sun_meta.get("priorYear"),
            "source": sun_meta.get("source"),
            "matched": with_sunshine,
        }
    with open(OUTPUT, "w") as f:
        json.dump(payload, f, indent=2)

    with_photo = sum(1 for m in mpps if m.get("photo"))
    with_riding = sum(1 for m in mpps if m.get("riding"))
    with_expenses = sum(1 for m in mpps if m.get("expenses"))
    with_urls = sum(1 for b in bills_meta if b.get("url"))
    parties = {}
    for m in mpps:
        parties[m["party"] or "?"] = parties.get(m["party"] or "?", 0) + 1

    print(f"Wrote {len(mpps)} MPPs ({with_photo} photos, {with_riding} with riding, "
          f"{with_expenses} with expenses, {with_sunshine} with sunshine pay)")
    print(f"Bill URLs: {with_urls}/{len(bills_meta)}")
    print(f"Parties: {parties}")
    print(f"Skipped empty/invalid rows: {skipped}")
    if sun_meta:
        unmatched = sun_meta.get("unmatched") or []
        print(f"Sunshine List {sun_meta.get('year')}: {with_sunshine} matched"
              + (f"; unmatched in fetch: {unmatched}" if unmatched else ""))
        if sunshine_notes:
            print(f"Sunshine vs sheet notes ({len(sunshine_notes)}):")
            for n in sunshine_notes[:20]:
                print(f"  • {n}")
            if len(sunshine_notes) > 20:
                print(f"  … +{len(sunshine_notes) - 20} more")

    # Sanity: Tyler Allsopp
    tyler = next((m for m in mpps if "Allsopp" in m["name"]), None)
    if tyler:
        exp = tyler.get("expenses") or {}
        print(f"Sample Tyler Allsopp: salary={tyler['salary']} riding={tyler['riding']!r} "
              f"photo={'yes' if tyler['photo'] else 'no'} votes={len(tyler['votes'])} "
              f"align={tyler['votingAlignment']}% expenses={exp.get('total')}")


if __name__ == "__main__":
    main()
